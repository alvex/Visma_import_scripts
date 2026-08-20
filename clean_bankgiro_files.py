from __future__ import annotations

import argparse
import re
import sys
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "edit"
TEST_FILE_NAME = "Bg819-5968_Insättningsuppgifter_20260415.xlsx"
FILE_PATTERN = "Bg*_Insättningsuppgifter_*.xlsx"
OUTPUT_PREFIX = "clean_bet_lista_"


@dataclass
class FileLog:
    file_name: str
    insattningsuppgift_found: bool = False
    belopp_insatt_found: bool = False
    insattningar_found: bool = False
    bankgiro_column_removed: bool = False
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def has_warnings(self) -> bool:
        return bool(self.warnings)

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)


def normalize_text(value: Any) -> str:
    """Normalize text for robust comparisons across Excel-export variants."""
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\u00a0", " ")
    text = text.replace("\r", " ")
    text = text.replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip().casefold()


def compact_normalized_text(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def find_cell_by_text(sheet: Worksheet, search_text: str):
    wanted = normalize_text(search_text)
    for row in sheet.iter_rows():
        for cell in row:
            if normalize_text(cell.value) == wanted:
                return cell
    return None


def find_cell_containing_text(sheet: Worksheet, search_text: str):
    wanted = normalize_text(search_text)
    for row in sheet.iter_rows():
        for cell in row:
            if wanted and wanted in normalize_text(cell.value):
                return cell
    return None


def find_section_header(sheet: Worksheet, search_text: str):
    exact = find_cell_by_text(sheet, search_text)
    if exact is not None:
        return exact
    return find_cell_containing_text(sheet, search_text)


def next_section_row(sheet: Worksheet, start_row: int, section_names: tuple[str, ...]) -> int:
    section_names_normalized = {normalize_text(name) for name in section_names}
    for row_idx in range(start_row + 1, sheet.max_row + 1):
        for cell in sheet[row_idx]:
            if normalize_text(cell.value) in section_names_normalized:
                return row_idx
    return sheet.max_row + 1


def clear_cell(cell) -> None:
    if isinstance(cell, MergedCell):
        return
    cell.value = None


def clear_insattningsuppgift_section(sheet: Worksheet) -> tuple[bool, list[str]]:
    warnings_found: list[str] = []
    section_cell = find_section_header(sheet, "Insättningsuppgift")
    if section_cell is None:
        return False, ['Sektionen "Insättningsuppgift" saknas.']

    end_row = next_section_row(sheet, section_cell.row, ("Belopp insatt på konto", "Insättningar"))
    datum_columns: set[int] = set()
    for row_idx in range(section_cell.row + 1, end_row):
        for cell in sheet[row_idx]:
            if normalize_text(cell.value) == "datum":
                datum_columns.add(cell.column)

    if not datum_columns:
        warnings_found.append('Kolumnrubriken "Datum" saknas i sektionen "Insättningsuppgift".')

    for row_idx in range(section_cell.row + 1, end_row):
        for cell in sheet[row_idx]:
            if cell.column not in datum_columns:
                clear_cell(cell)

    return True, warnings_found


def clean_belopp_insatt_section(sheet: Worksheet) -> tuple[bool, list[str]]:
    warnings_found: list[str] = []
    section_cell = find_section_header(sheet, "Belopp insatt på konto")
    if section_cell is None:
        return False, ['Sektionen "Belopp insatt på konto" saknas.']

    end_row = next_section_row(sheet, section_cell.row, ("Insättningar",))
    keep_columns: set[int] = set()

    for row_idx in range(section_cell.row + 1, end_row):
        for cell in sheet[row_idx]:
            text = normalize_text(cell.value)
            if text == "belopp" or text.startswith("löpnummer"):
                keep_columns.add(cell.column)

    if not any(normalize_text(sheet.cell(row=row_idx, column=col).value) == "belopp"
               for row_idx in range(section_cell.row + 1, end_row)
               for col in keep_columns):
        warnings_found.append('Kolumnrubriken "Belopp" saknas i sektionen "Belopp insatt på konto".')

    if not any(normalize_text(sheet.cell(row=row_idx, column=col).value).startswith("löpnummer")
               for row_idx in range(section_cell.row + 1, end_row)
               for col in keep_columns):
        warnings_found.append('Rubriken "Löpnummer" saknas i sektionen "Belopp insatt på konto".')

    for row_idx in range(section_cell.row + 1, end_row):
        for cell in sheet[row_idx]:
            if cell.column not in keep_columns:
                clear_cell(cell)

    return True, warnings_found


def is_bankgiro_avinummer_header(value: Any) -> bool:
    text = compact_normalized_text(value)
    return text in {"bankgironummer/avinummer", "bankgironummeravinummer"}


def remove_bankgiro_avinummer_column(sheet: Worksheet) -> tuple[bool, bool, list[str]]:
    warnings_found: list[str] = []
    section_cell = find_section_header(sheet, "Insättningar")
    if section_cell is None:
        return False, False, ['Sektionen "Insättningar" saknas.']

    for row_idx in range(section_cell.row + 1, sheet.max_row + 1):
        for cell in sheet[row_idx]:
            if is_bankgiro_avinummer_header(cell.value):
                sheet.delete_cols(cell.column, 1)
                return True, True, warnings_found

    warnings_found.append('Kolumnen "Bankgironummer/Avinummer" saknas i sektionen "Insättningar".')
    return True, False, warnings_found


def process_workbook(file_path: Path, output_dir: Path) -> FileLog:
    log = FileLog(file_name=file_path.name)

    try:
        with warnings.catch_warnings(record=True) as caught_warnings:
            warnings.simplefilter("always")
            workbook = load_workbook(file_path)
            for warning in caught_warnings:
                if "Print area cannot be set" in str(warning.message):
                    continue
                log.warnings.append(str(warning.message))

        for sheet in workbook.worksheets:
            found, section_warnings = clear_insattningsuppgift_section(sheet)
            log.insattningsuppgift_found = log.insattningsuppgift_found or found
            log.warnings.extend(section_warnings)

            found, section_warnings = clean_belopp_insatt_section(sheet)
            log.belopp_insatt_found = log.belopp_insatt_found or found
            log.warnings.extend(section_warnings)

            found, removed, section_warnings = remove_bankgiro_avinummer_column(sheet)
            log.insattningar_found = log.insattningar_found or found
            log.bankgiro_column_removed = log.bankgiro_column_removed or removed
            log.warnings.extend(section_warnings)

        output_dir.mkdir(parents=True, exist_ok=True)
        output_name = f"{OUTPUT_PREFIX}{file_path.name}"
        workbook.save(output_dir / output_name)

    except Exception as exc:  # noqa: BLE001 - top-level file processing should report and continue.
        log.errors.append(str(exc))

    return log


def print_file_log(log: FileLog) -> None:
    print(f"\nFil: {log.file_name}")
    print(f'  Insättningsuppgift hittades: {"ja" if log.insattningsuppgift_found else "nej"}')
    print(f'  Belopp insatt på konto hittades: {"ja" if log.belopp_insatt_found else "nej"}')
    print(f'  Insättningar hittades: {"ja" if log.insattningar_found else "nej"}')
    print(f'  Bankgironummer/Avinummer borttagen: {"ja" if log.bankgiro_column_removed else "nej"}')

    for warning_text in log.warnings:
        print(f"  VARNING: {warning_text}")
    for error_text in log.errors:
        print(f"  FEL: {error_text}")


def print_summary(logs: list[FileLog], found_count: int, output_dir: Path) -> None:
    processed_count = sum(1 for log in logs if not log.has_errors)
    warning_count = sum(1 for log in logs if log.has_warnings)
    error_count = sum(1 for log in logs if log.has_errors)

    print("\nSammanfattning")
    print(f"  Antal hittade filer: {found_count}")
    print(f"  Antal bearbetade filer: {processed_count}")
    print(f"  Antal filer med varningar: {warning_count}")
    print(f"  Antal filer med fel: {error_count}")
    print(f"  Sökväg till sparade filer: {output_dir}")


def process_single_test_file(output_dir: Path) -> int:
    test_file = BASE_DIR / TEST_FILE_NAME
    if not test_file.exists():
        print(f"FEL: Testfilen finns inte: {test_file}", file=sys.stderr)
        return 1

    log = process_workbook(test_file, output_dir)
    print_file_log(log)
    print_summary([log], found_count=1, output_dir=output_dir)

    if log.has_errors:
        return 1

    print(
        f"\nTestfil skapad i {output_dir}. Kontrollera filen manuellt. "
        "Om resultatet är godkänt, kör scriptet igen med --all för att bearbeta alla filer."
    )
    return 0


def process_all_files(input_dir: Path, output_dir: Path) -> int:
    if not input_dir.is_dir():
        print(f"FEL: Indatamappen finns inte: {input_dir}", file=sys.stderr)
        return 1

    resolved_output = output_dir.resolve()
    files = sorted(
        path
        for path in input_dir.glob(FILE_PATTERN)
        if path.is_file()
        and not path.name.startswith("~$")
        and path.resolve().parent != resolved_output
    )
    logs: list[FileLog] = []

    if not files:
        print(f'VARNING: Inga filer matchade "{FILE_PATTERN}" i {input_dir}.')

    for file_path in files:
        log = process_workbook(file_path, output_dir)
        logs.append(log)
        print_file_log(log)

    print_summary(logs, found_count=len(files), output_dir=output_dir)
    return 1 if any(log.has_errors for log in logs) else 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Skapa rensade kopior av Bankgiro-filer utan att ändra originalfiler."
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help=f'Bearbeta alla filer som matchar "{FILE_PATTERN}". Utan flaggan bearbetas bara testfilen.',
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help=f'Mapp att läsa filer från (matchar "{FILE_PATTERN}"). Standard: {BASE_DIR}.',
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help=f"Mapp att spara rensade filer i. Standard: {OUTPUT_DIR}.",
    )
    args = parser.parse_args()

    output_dir = args.output if args.output is not None else OUTPUT_DIR

    # Att ange --input eller --output innebär batch-körning över hela indatamappen.
    if args.input is not None or args.output is not None or args.all:
        input_dir = args.input if args.input is not None else BASE_DIR
        return process_all_files(input_dir, output_dir)

    return process_single_test_file(output_dir)


if __name__ == "__main__":
    raise SystemExit(main())
