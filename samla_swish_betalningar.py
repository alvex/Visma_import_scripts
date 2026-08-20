"""Samla rensade Swish-filer till en gemensam Excel-fil.

Läser alla filer som matchar "clean_swish_lista_*.xlsx" och skapar en
sammanställning samla_swish_bet_lista_<dagens datum>.xlsx med alla
Swish-transaktioner. Period och filtotal visas en gång per källfil.

Eget separat steg - rör inte Bankgiro-flödet.

Körs via dashboarden med --input/--output, eller fristående:
    python samla_swish_betalningar.py --input "C:\\...\\output" --output "C:\\...\\output"
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

BASE_DIR = Path(__file__).resolve().parent
FILE_PATTERN = "clean_swish_lista_*.xlsx"
OUTPUT_PREFIX = "samla_swish_bet_lista"

# Kolumner som läses ur varje rensad fil.
SOURCE_COLUMNS = ["Bokförd", "Avsändare", "Meddelande", "Insättningar"]
# Kolumner i resultatfilen.
OUTPUT_COLUMNS = ["Bokförd", "Avsändare", "Meddelande", "Insättningar", "Period", "Filtotal"]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace(" ", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def header_key(value: Any) -> str:
    return normalize_text(value).casefold()


def to_decimal(value: Any) -> Decimal | None:
    """Tolka belopp till Decimal. Hanterar tal och svenska kommasträngar."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = normalize_text(value).replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def find_period(sheet: Worksheet) -> str:
    for row in sheet.iter_rows():
        for cell in row:
            text = normalize_text(cell.value)
            if text.lower().startswith("period:"):
                return text.split(":", 1)[1].strip()
    return ""


def find_header_row(sheet: Worksheet) -> tuple[int | None, dict[str, int]]:
    """Hitta raden som innehåller källkolumnerna. Returnerar (rad, {namn: kolindex})."""
    wanted = {header_key(name): name for name in SOURCE_COLUMNS}
    for row_idx in range(1, sheet.max_row + 1):
        found: dict[str, int] = {}
        for cell in sheet[row_idx]:
            key = header_key(cell.value)
            if key in wanted:
                found[wanted[key]] = cell.column
        if all(name in found for name in SOURCE_COLUMNS):
            return row_idx, found
    return None, {}


def collect_file(path: Path) -> tuple[str, list[dict[str, Any]]] | None:
    """Läs en rensad Swish-fil. Returnerar (period, transaktionsrader) eller None vid fel.

    Transaktionsrader = rader med ett giltigt belopp i Insättningar. Totalraden
    (där bara Summa är ifylld) saknar Insättningar och hoppas därför över.
    """
    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001 - logga och fortsätt med nästa fil.
        print(f"  FEL: Kunde inte öppna filen: {exc}", file=sys.stderr)
        return None

    sheet = workbook.worksheets[0]
    period = find_period(sheet)
    header_row, col = find_header_row(sheet)
    if header_row is None:
        print(f'  FEL: Hittade ingen tabell med kolumnerna {", ".join(SOURCE_COLUMNS)}.', file=sys.stderr)
        return None

    rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        belopp = to_decimal(sheet.cell(row=row_idx, column=col["Insättningar"]).value)
        if belopp is None:
            continue  # totalrad eller tom rad
        rows.append({
            "bokford": normalize_text(sheet.cell(row=row_idx, column=col["Bokförd"]).value),
            "avsandare": normalize_text(sheet.cell(row=row_idx, column=col["Avsändare"]).value),
            "meddelande": normalize_text(sheet.cell(row=row_idx, column=col["Meddelande"]).value),
            "belopp": belopp,
        })
    return period, rows


def unique_output_path(output_dir: Path) -> Path:
    today = date.today().strftime("%Y-%m-%d")
    base = output_dir / f"{OUTPUT_PREFIX}_{today}.xlsx"
    if not base.exists():
        return base
    counter = 2
    while True:
        candidate = output_dir / f"{OUTPUT_PREFIX}_{today}_{counter}.xlsx"
        if not candidate.exists():
            return candidate
        counter += 1


def export_summary(groups: list[dict[str, Any]], output_dir: Path) -> tuple[Path, Decimal]:
    output_path = unique_output_path(output_dir)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Swish"
    sheet.append(OUTPUT_COLUMNS)

    grand_total = Decimal("0.00")
    for group in groups:
        rows = group["rows"]
        period = group["period"]
        filtotal = sum((row["belopp"] for row in rows), Decimal("0.00"))
        grand_total += filtotal
        for index, row in enumerate(rows):
            first = index == 0
            sheet.append([
                row["bokford"],
                row["avsandare"],
                row["meddelande"],
                float(row["belopp"]),
                period if first else None,
                float(filtotal) if first else None,
            ])

    summary_row = sheet.max_row + 1
    sheet.cell(row=summary_row, column=1, value="SUMMA")
    sheet.cell(row=summary_row, column=4, value=float(grand_total))

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for cell in sheet[summary_row]:
        cell.font = Font(bold=True)

    for row_idx in range(2, summary_row + 1):
        sheet.cell(row=row_idx, column=4).number_format = "#,##0.00"
        sheet.cell(row=row_idx, column=6).number_format = "#,##0.00"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:F{summary_row}"

    widths = {"A": 12, "B": 30, "C": 42, "D": 14, "E": 24, "F": 14}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path, grand_total


def find_matching_files(input_dir: Path) -> list[Path]:
    return sorted(
        (path for path in input_dir.glob(FILE_PATTERN)
         if path.is_file() and not path.name.startswith("~$")),
        key=lambda path: path.name.casefold(),
    )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Samla rensade Swish-filer (clean_swish_lista_*.xlsx) till en gemensam fil."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help=f'Mapp att läsa rensade Swish-filer från ("{FILE_PATTERN}"). Standard: {BASE_DIR}.',
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Mapp att spara sammanställningen i. Standard: samma som --input.",
    )
    args = parser.parse_args()

    input_dir = args.input if args.input is not None else BASE_DIR
    output_dir = args.output if args.output is not None else input_dir

    if not input_dir.is_dir():
        print(f"FEL: Indatamappen finns inte: {input_dir}", file=sys.stderr)
        return 1
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:  # noqa: BLE001
        print(f"FEL: Kunde inte skapa utdatamappen: {exc}", file=sys.stderr)
        return 1

    files = find_matching_files(input_dir)
    if not files:
        print(f'FEL: Inga filer matchade "{FILE_PATTERN}" i {input_dir}.', file=sys.stderr)
        return 1

    print(f"Filer som hittades ({len(files)}):")
    for path in files:
        print(f"  - {path.name}")

    groups: list[dict[str, Any]] = []
    failed = 0
    for path in files:
        print(f"\nBehandlar: {path.name}")
        result = collect_file(path)
        if result is None:
            failed += 1
            continue
        period, rows = result
        print(f"  Period: {period or '(saknas)'}")
        print(f"  Transaktioner: {len(rows)}")
        if rows:
            groups.append({"source_file": path.name, "period": period, "rows": rows})

    total_rows = sum(len(group["rows"]) for group in groups)
    if total_rows == 0:
        print("\nInga transaktioner hittades. Ingen output-fil skapades.", file=sys.stderr)
        return 1

    output_path, grand_total = export_summary(groups, output_dir)

    print("\nSammanställning klar.")
    print(f"  Filer hittade: {len(files)}")
    print(f"  Filer med fel: {failed}")
    print(f"  Sammanställda transaktioner: {total_rows}")
    print(f"  Total summa: {grand_total}")
    print(f"  Skapad fil: {output_path}")

    return 0 if failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
