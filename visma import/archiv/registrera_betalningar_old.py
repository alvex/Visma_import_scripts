from __future__ import annotations

import argparse
import csv
import re
import sys
import time
from dataclasses import dataclass, asdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from playwright.sync_api import BrowserContext, Page, TimeoutError as PlaywrightTimeoutError, sync_playwright


BASE_URL = "https://hemfresh.com/fakturor/index.php"
PAYMENT_URL_TEMPLATE = (
    "https://hemfresh.com/fakturor/index.php"
    "?module=payments&view=process&invoice={invoice}&op=pay_selected_invoice"
)
SESSION_FILE = Path("hemfresh_session.json")
LOG_DIR = Path("payment_logs")
SCREENSHOT_DIR = Path("screenshots")

REQUIRED_COLUMNS = ["Datum", "Avsändare", "Betalningsreferens", "Belopp"]
MIN_DELAY_SECONDS = 1.0
MAX_DELAY_SECONDS = 2.0


@dataclass
class PaymentInput:
    row_number: int
    datum: date
    avsandare: str
    betalningsreferens: str
    belopp: Decimal
    fakturanummer: str


@dataclass
class PaymentLog:
    Datum: str
    Avsändare: str
    Betalningsreferens: str
    Fakturanummer: str
    Belopp_fran_Excel: str
    Belopp_pa_fakturasidan: str
    Status: str
    Meddelande: str
    Tidpunkt: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_amount(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, int | float):
        return Decimal(str(value)).quantize(Decimal("0.01"))

    text = normalize_text(value)
    text = re.sub(r"(?i)\b(sek|kr|kronor)\b", "", text)
    text = re.sub(r"[^0-9,.\- ]", "", text).replace(" ", "")
    if not text or text in {"-", ".", ","}:
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    elif text.count(".") > 1:
        parts = text.split(".")
        text = "".join(parts[:-1]) + "." + parts[-1]

    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def normalize_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None

    text = normalize_text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def amount_for_log(amount: Decimal | None) -> str:
    if amount is None:
        return ""
    return f"{amount:.2f}"


def ask_excel_path() -> Path | None:
    raw = input("Ange sökvägen till Excel-filen med betalningar: ").strip().strip('"').strip("'")
    path = Path(raw)
    if not path.exists() or not path.is_file():
        print(f"Fel: Excel-filen finns inte: {path}", file=sys.stderr)
        return None
    return path


def ask_run_mode() -> bool:
    print("\nVälj körläge:")
    print("  1. dry-run/testläge (registrerar inget)")
    print("  2. riktig registrering")
    choice = input("Ange 1 eller 2: ").strip()
    if choice == "2":
        confirmation = input('Skriv REGISTRERA för att bekräfta riktig registrering: ').strip()
        if confirmation == "REGISTRERA":
            return False
        print("Bekräftelsen matchade inte. Körningen fortsätter i dry-run/testläge.")
        return True
    return True


def column_key(value: Any) -> str:
    return normalize_text(value).casefold()


def read_payments_from_excel(path: Path) -> tuple[list[PaymentInput], list[PaymentLog]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active

    header_row = None
    column_map: dict[str, int] = {}
    required = {column_key(name): name for name in REQUIRED_COLUMNS}

    for row_idx in range(1, sheet.max_row + 1):
        found: dict[str, int] = {}
        for cell in sheet[row_idx]:
            key = column_key(cell.value)
            if key in required:
                found[required[key]] = cell.column
        if all(name in found for name in REQUIRED_COLUMNS):
            header_row = row_idx
            column_map = found
            break

    if header_row is None:
        raise ValueError(f"Excel-filen saknar någon av kolumnerna: {', '.join(REQUIRED_COLUMNS)}")

    payments: list[PaymentInput] = []
    skipped_logs: list[PaymentLog] = []

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        values = {
            name: sheet.cell(row=row_idx, column=column_map[name]).value
            for name in REQUIRED_COLUMNS
        }
        if all(value in (None, "") for value in values.values()):
            continue

        datum = normalize_date(values["Datum"])
        amount = normalize_amount(values["Belopp"])
        reference = normalize_text(values["Betalningsreferens"])
        invoice = extract_invoice_number(reference)

        if datum is None or amount is None or invoice is None:
            skipped_logs.append(
                make_log(
                    datum=datum.isoformat() if datum else normalize_text(values["Datum"]),
                    avsandare=normalize_text(values["Avsändare"]),
                    reference=reference,
                    invoice=invoice or "",
                    excel_amount=amount,
                    page_amount=None,
                    status="HOPPAD",
                    message="Raden saknar säkert datum, belopp eller fakturanummer.",
                )
            )
            continue

        payments.append(
            PaymentInput(
                row_number=row_idx,
                datum=datum,
                avsandare=normalize_text(values["Avsändare"]),
                betalningsreferens=reference,
                belopp=amount,
                fakturanummer=invoice,
            )
        )

    return payments, skipped_logs


def extract_invoice_number(reference: str) -> str | None:
    text = normalize_text(reference)
    if not text:
        return None

    keyword_patterns = [
        r"(?i)\bfaktura\s*(?:nr|nummer|#|:)?\s*(\d{4,7})\b",
        r"(?i)\bfak(?:t|tnr|\.|turanr)?\s*(?:nr|nummer|#|:)?\s*(\d{4,7})\b",
        r"(?i)\binvoice\s*(?:no|nr|#|:)?\s*(\d{4,7})\b",
    ]
    for pattern in keyword_patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)

    candidates = re.findall(r"(?<!\d)(\d{4,7})(?!\d)", text)
    candidates = [candidate for candidate in candidates if not candidate.startswith(("19", "20"))]
    unique_candidates = sorted(set(candidates))
    if len(unique_candidates) == 1:
        return unique_candidates[0]
    return None


def make_log(
    datum: str,
    avsandare: str,
    reference: str,
    invoice: str,
    excel_amount: Decimal | None,
    page_amount: Decimal | None,
    status: str,
    message: str,
) -> PaymentLog:
    return PaymentLog(
        Datum=datum,
        Avsändare=avsandare,
        Betalningsreferens=reference,
        Fakturanummer=invoice,
        Belopp_fran_Excel=amount_for_log(excel_amount),
        Belopp_pa_fakturasidan=amount_for_log(page_amount),
        Status=status,
        Meddelande=message,
        Tidpunkt=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )


def ensure_login(context: BrowserContext) -> Page:
    page = context.new_page()
    page.goto(BASE_URL, wait_until="domcontentloaded")
    print("\nWebbläsaren är öppen.")
    print("Om du inte är inloggad: logga in manuellt i webbläsaren.")
    input("Tryck Enter här när du ser fakturasystemet och är inloggad...")
    context.storage_state(path=str(SESSION_FILE))
    print(f"Session sparad i {SESSION_FILE}.")
    return page


def extract_att_betala_amount(page: Page) -> Decimal | None:
    body_text = page.locator("body").inner_text(timeout=10000)
    match = re.search(r"Att\s+betala\s*:?\s*([\-0-9\s.,]+(?:kr|SEK)?)", body_text, flags=re.IGNORECASE)
    if not match:
        return None
    return normalize_amount(match.group(1))


def page_looks_paid(page: Page, page_amount: Decimal | None) -> bool:
    body_text = page.locator("body").inner_text(timeout=10000).casefold()
    paid_words = ["redan betald", "är betald", "betald faktura", "att betala: 0"]
    return page_amount == Decimal("0.00") or any(word in body_text for word in paid_words)


def try_fill(page: Page, selectors: list[str], value: str) -> bool:
    for selector in selectors:
        try:
            locator = page.locator(selector).first
            if locator.count() > 0 and locator.is_visible(timeout=1000):
                locator.fill(value)
                return True
        except PlaywrightTimeoutError:
            continue
        except Exception:
            continue
    return False


def fill_payment_form(page: Page, payment: PaymentInput) -> tuple[bool, str]:
    date_value = payment.datum.isoformat()
    amount_value = f"{payment.belopp:.2f}"

    date_ok = try_fill(
        page,
        [
            'input[name*="date" i]',
            'input[id*="date" i]',
            'input[name*="datum" i]',
            'input[id*="datum" i]',
            'input[type="date"]',
        ],
        date_value,
    )
    amount_ok = try_fill(
        page,
        [
            'input[name*="amount" i]',
            'input[id*="amount" i]',
            'input[name*="belopp" i]',
            'input[id*="belopp" i]',
            'input[name*="paid" i]',
            'input[id*="paid" i]',
        ],
        amount_value,
    )

    if not date_ok or not amount_ok:
        return False, f"Kunde inte fylla formuläret säkert. Datumfält: {date_ok}, beloppsfält: {amount_ok}."
    return True, "Formuläret fylldes i."


def click_register_button(page: Page) -> bool:
    button_selectors = [
        'button:has-text("Registrera")',
        'button:has-text("Spara")',
        'button:has-text("Betala")',
        'input[type="submit"][value*="Registrera" i]',
        'input[type="submit"][value*="Spara" i]',
        'input[type="submit"][value*="Betala" i]',
    ]
    for selector in button_selectors:
        try:
            locator = page.locator(selector).first
            if locator.count() > 0 and locator.is_visible(timeout=1000):
                locator.click()
                return True
        except Exception:
            continue
    return False


def save_error_screenshot(page: Page, invoice: str) -> Path | None:
    try:
        SCREENSHOT_DIR.mkdir(exist_ok=True)
        path = SCREENSHOT_DIR / f"fel_invoice_{invoice}_{datetime.now():%Y%m%d_%H%M%S}.png"
        page.screenshot(path=str(path), full_page=True)
        return path
    except Exception:
        return None


def process_payment(page: Page, payment: PaymentInput, dry_run: bool) -> PaymentLog:
    page_amount: Decimal | None = None
    url = PAYMENT_URL_TEMPLATE.format(invoice=payment.fakturanummer)

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        page_amount = extract_att_betala_amount(page)

        if page_amount is None:
            screenshot = save_error_screenshot(page, payment.fakturanummer)
            message = 'Kunde inte hitta texten "Att betala:" eller tolka beloppet.'
            if screenshot:
                message += f" Screenshot: {screenshot}"
            return make_payment_log(payment, page_amount, "FEL", message)

        if page_looks_paid(page, page_amount):
            return make_payment_log(payment, page_amount, "HOPPAD", "Fakturan verkar redan vara betald.")

        if page_amount != payment.belopp:
            screenshot = save_error_screenshot(page, payment.fakturanummer)
            message = "Beloppet matchar inte exakt. Ingen registrering gjordes."
            if screenshot:
                message += f" Screenshot: {screenshot}"
            return make_payment_log(payment, page_amount, "FEL", message)

        form_ok, form_message = fill_payment_form(page, payment)
        if not form_ok:
            screenshot = save_error_screenshot(page, payment.fakturanummer)
            if screenshot:
                form_message += f" Screenshot: {screenshot}"
            return make_payment_log(payment, page_amount, "FEL", form_message)

        if dry_run:
            return make_payment_log(payment, page_amount, "DRY-RUN", "Belopp matchade. Skulle registrera betalningen.")

        if not click_register_button(page):
            screenshot = save_error_screenshot(page, payment.fakturanummer)
            message = "Kunde inte hitta en säker knapp för att registrera betalningen."
            if screenshot:
                message += f" Screenshot: {screenshot}"
            return make_payment_log(payment, page_amount, "FEL", message)

        page.wait_for_load_state("domcontentloaded", timeout=15000)
        if page_looks_paid(page, Decimal("0.00")):
            return make_payment_log(payment, page_amount, "REGISTRERAD", "Betalningen verkar vara registrerad.")

        return make_payment_log(payment, page_amount, "OKÄND", "Klick utfört, men bekräftelse kunde inte verifieras säkert.")

    except Exception as exc:  # noqa: BLE001 - per-row errors should be logged and processing should continue.
        try:
            screenshot = save_error_screenshot(page, payment.fakturanummer)
        except Exception:
            screenshot = None
        message = f"Oväntat fel: {exc}"
        if screenshot:
            message += f" Screenshot: {screenshot}"
        return make_payment_log(payment, page_amount, "FEL", message)


def make_payment_log(payment: PaymentInput, page_amount: Decimal | None, status: str, message: str) -> PaymentLog:
    return make_log(
        datum=payment.datum.isoformat(),
        avsandare=payment.avsandare,
        reference=payment.betalningsreferens,
        invoice=payment.fakturanummer,
        excel_amount=payment.belopp,
        page_amount=page_amount,
        status=status,
        message=message,
    )


def write_logs(logs: list[PaymentLog]) -> Path:
    LOG_DIR.mkdir(exist_ok=True)
    log_path = LOG_DIR / f"betalningsregistrering_{datetime.now():%Y%m%d_%H%M%S}.csv"
    with log_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(asdict(logs[0]).keys()), delimiter=";")
        writer.writeheader()
        for row in logs:
            writer.writerow(asdict(row))
    return log_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Registrera fakturabetalningar via Playwright.")
    parser.add_argument("--all", action="store_true", help="Behandla alla giltiga rader. Utan flaggan behandlas bara första raden.")
    args = parser.parse_args()

    excel_path = ask_excel_path()
    if excel_path is None:
        return 1

    dry_run = ask_run_mode()
    if not dry_run and not args.all:
        print("Riktig registrering kräver --all eller en medveten körning. Kör om med --all när dry-run är kontrollerad.")
        return 1

    payments, logs = read_payments_from_excel(excel_path)
    if not payments:
        print("Inga giltiga betalningsrader hittades.")
        if logs:
            print(f"Loggfil: {write_logs(logs)}")
        return 1

    payments_to_process = payments if args.all else payments[:1]
    print(f"\nGiltiga rader i Excel: {len(payments)}")
    print(f"Rader som behandlas nu: {len(payments_to_process)}")
    print(f"Körläge: {'dry-run/testläge' if dry_run else 'riktig registrering'}")

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False)
        context_kwargs = {"storage_state": str(SESSION_FILE)} if SESSION_FILE.exists() else {}
        context = browser.new_context(**context_kwargs)
        page = ensure_login(context)

        for index, payment in enumerate(payments_to_process, start=1):
            print(f"\n[{index}/{len(payments_to_process)}] Faktura {payment.fakturanummer}, belopp {payment.belopp:.2f}")
            log = process_payment(page, payment, dry_run=dry_run)
            logs.append(log)
            print(f"  {log.Status}: {log.Meddelande}")
            time.sleep(MIN_DELAY_SECONDS if index % 2 else MAX_DELAY_SECONDS)

        context.storage_state(path=str(SESSION_FILE))
        browser.close()

    log_path = write_logs(logs)
    print("\nKörning klar.")
    print(f"Loggfil: {log_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
