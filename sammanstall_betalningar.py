from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


BASE_DIR = Path(__file__).resolve().parent
FILE_PATTERN = "clean_bet_lista_*.xlsx"
OUTPUT_PREFIX = "samlade_betalningar"
OUTPUT_COLUMNS = ["Datum", "Avsändare", "Betalningsreferens", "Belopp", "Löpnummer", "Total"]

DATE_RE = re.compile(r"(?<!\d)(\d{4}[-./]?\d{2}[-./]?\d{2}|\d{2}[-./]\d{2}[-./]\d{4})(?!\d)")
FILENAME_DATE_RE = re.compile(r"_(\d{8})(?:\.xlsx)?$", re.IGNORECASE)


@dataclass
class PaymentRow:
    source_file: str
    datum: date
    avsandare: str
    betalningsreferens: str
    belopp: Decimal
    lopnummer: str
    total: Decimal | None


@dataclass
class FileResult:
    path: Path
    rows: list[PaymentRow] = field(default_factory=list)
    datum: date | None = None
    total: Decimal | None = None
    lopnummer: str = ""
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def has_warnings(self) -> bool:
        return bool(self.warnings)

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\u00a0", " ")
    text = text.replace("\r", " ")
    text = text.replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip().casefold()


def compact_text(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def normalize_amount(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, int | float):
        return Decimal(str(value)).quantize(Decimal("0.01"))

    text = str(value).strip()
    if not text:
        return None

    text = text.replace("\u00a0", " ")
    text = re.sub(r"(?i)\b(sek|kr)\b", "", text)
    text = re.sub(r"[^0-9,.\- ]", "", text).strip()
    text = text.replace(" ", "")
    if not text or text in {"-", ".", ","}:
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    elif text.count(".") > 1:
        parts = text.split(".")
        text = "".join(parts[:-1]) + "." + parts[-1]

    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def normalize_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None

    text = str(value)
    match = DATE_RE.search(text)
    if not match:
        return None

    token = match.group(1)
    formats = ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d", "%d-%m-%Y", "%d.%m.%Y", "%d/%m/%Y")
    for fmt in formats:
        try:
            return datetime.strptime(token, fmt).date()
        except ValueError:
            continue
    return None


def find_matching_files(folder_path: Path) -> list[Path]:
    return sorted(
        (
            path
            for path in folder_path.glob(FILE_PATTERN)
            if path.is_file() and not path.name.startswith("~$")
        ),
        key=lambda item: item.name.casefold(),
    )


def load_workbook_safe(file_path: Path):
    try:
        return load_workbook(file_path, data_only=True)
    except Exception as exc:  # noqa: BLE001 - robust batch processing should log and continue.
        return exc


def find_cell_by_text(sheet: Worksheet, search_text: str, contains: bool = False):
    wanted = normalize_text(search_text)
    for row in sheet.iter_rows():
        for cell in row:
            text = normalize_text(cell.value)
            if text == wanted or (contains and wanted in text):
                return cell
    return None


def next_section_row(sheet: Worksheet, start_row: int, section_names: tuple[str, ...]) -> int:
    wanted = {normalize_text(name) for name in section_names}
    for row_idx in range(start_row + 1, sheet.max_row + 1):
        for cell in sheet[row_idx]:
            if normalize_text(cell.value) in wanted:
                return row_idx
    return sheet.max_row + 1


def find_date(sheet: Worksheet, file_path: Path) -> date | None:
    section_cell = find_cell_by_text(sheet, "Insättningsuppgift", contains=True)
    if section_cell is not None:
        end_row = next_section_row(sheet, section_cell.row, ("Belopp insatt på konto", "Insättningar"))
        for row_idx in range(section_cell.row, end_row):
            for cell in sheet[row_idx]:
                if normalize_text(cell.value) == "datum":
                    for nearby_row in range(row_idx, min(end_row, row_idx + 4)):
                        parsed = normalize_date(sheet.cell(row=nearby_row, column=cell.column).value)
                        if parsed:
                            return parsed

        for row_idx in range(section_cell.row, end_row):
            for cell in sheet[row_idx]:
                parsed = normalize_date(cell.value)
                if parsed:
                    return parsed

    for row in sheet.iter_rows():
        for cell in row:
            parsed = normalize_date(cell.value)
            if parsed:
                return parsed

    filename_match = FILENAME_DATE_RE.search(file_path.name)
    if filename_match:
        return normalize_date(filename_match.group(1))

    return None


def extract_lopnummer(value: Any) -> str | None:
    text = str(value or "")
    match = re.search(r"löpnummer\s*[:#-]?\s*(\S+)", text, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return None


def find_deposit_total_and_lopnummer(sheet: Worksheet) -> tuple[Decimal | None, str]:
    section_cell = find_cell_by_text(sheet, "Belopp insatt på konto", contains=True)
    if section_cell is None:
        return None, ""

    end_row = next_section_row(sheet, section_cell.row, ("Insättningar",))
    lopnummer = ""
    belopp_column: int | None = None
    total_candidates: list[Decimal] = []

    for row_idx in range(section_cell.row, end_row):
        row_values = [cell.value for cell in sheet[row_idx]]
        row_text = " ".join(str(value) for value in row_values if value is not None)
        found_lopnummer = extract_lopnummer(row_text)
        if found_lopnummer:
            lopnummer = found_lopnummer

        for cell in sheet[row_idx]:
            if normalize_text(cell.value) == "belopp":
                belopp_column = cell.column

    for row_idx in range(section_cell.row + 1, end_row):
        row_text = " ".join(normalize_text(cell.value) for cell in sheet[row_idx] if cell.value is not None)
        row_amounts = [normalize_amount(cell.value) for cell in sheet[row_idx]]
        row_amounts = [amount for amount in row_amounts if amount is not None]

        if "totalt" in row_text and row_amounts:
            total_candidates.extend(row_amounts)
        elif belopp_column is not None:
            amount = normalize_amount(sheet.cell(row=row_idx, column=belopp_column).value)
            if amount is not None:
                total_candidates.append(amount)
        else:
            total_candidates.extend(row_amounts)

    total = total_candidates[-1] if total_candidates else None
    return total, lopnummer


def header_matches(value: Any, expected: str) -> bool:
    return compact_text(value) == compact_text(expected)


def find_payment_table(sheet: Worksheet) -> tuple[int | None, dict[str, int]]:
    required = ("Avsändare", "Betalningsreferens", "Belopp")

    for row_idx in range(1, sheet.max_row + 1):
        columns: dict[str, int] = {}
        for cell in sheet[row_idx]:
            for header in required:
                if header_matches(cell.value, header):
                    columns[header] = cell.column

        if all(header in columns for header in required):
            return row_idx, columns

    return None, {}


def row_is_empty(sheet: Worksheet, row_idx: int, columns: dict[str, int]) -> bool:
    return all(sheet.cell(row=row_idx, column=col_idx).value in (None, "") for col_idx in columns.values())


def looks_like_new_section(sheet: Worksheet, row_idx: int, columns: dict[str, int]) -> bool:
    populated = [cell for cell in sheet[row_idx] if cell.value not in (None, "")]
    if len(populated) != 1:
        return False

    only_cell = populated[0]
    if only_cell.column in columns.values():
        return False

    return bool(normalize_text(only_cell.value))


def collect_data_from_file(file_path: Path) -> FileResult:
    result = FileResult(path=file_path)
    workbook_or_error = load_workbook_safe(file_path)
    if isinstance(workbook_or_error, Exception):
        result.errors.append(f"Kunde inte öppna filen: {workbook_or_error}")
        return result

    workbook = workbook_or_error
    for sheet in workbook.worksheets:
        datum = find_date(sheet, file_path)
        total, lopnummer = find_deposit_total_and_lopnummer(sheet)
        header_row, columns = find_payment_table(sheet)

        result.datum = result.datum or datum
        result.total = result.total if result.total is not None else total
        result.lopnummer = result.lopnummer or lopnummer

        if header_row is None:
            continue

        if datum is None:
            continue

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            if row_is_empty(sheet, row_idx, columns):
                break
            if looks_like_new_section(sheet, row_idx, columns):
                break

            avsandare = sheet.cell(row=row_idx, column=columns["Avsändare"]).value
            referens = sheet.cell(row=row_idx, column=columns["Betalningsreferens"]).value
            belopp = normalize_amount(sheet.cell(row=row_idx, column=columns["Belopp"]).value)

            if belopp is None:
                continue
            if avsandare in (None, "") and referens in (None, ""):
                continue

            result.rows.append(
                PaymentRow(
                    source_file=file_path.name,
                    datum=datum,
                    avsandare=str(avsandare or ""),
                    betalningsreferens=str(referens or ""),
                    belopp=belopp,
                    lopnummer=lopnummer,
                    total=total,
                )
            )

        break

    if result.datum is None:
        result.warnings.append("Datum saknas eller kunde inte tolkas.")
    if result.total is None:
        result.warnings.append('Belopp insatt på konto saknas eller kunde inte tolkas.')
    if not result.lopnummer:
        result.warnings.append("Löpnummer saknas.")
    if not result.rows:
        result.warnings.append("Tabellen med Avsändare / Betalningsreferens / Belopp saknas eller gav inga giltiga rader.")

    return result


def unique_output_path(folder_path: Path) -> Path:
    today = date.today().strftime("%Y-%m-%d")
    base_path = folder_path / f"{OUTPUT_PREFIX}_{today}.xlsx"
    if not base_path.exists():
        return base_path

    counter = 2
    while True:
        candidate = folder_path / f"{OUTPUT_PREFIX}_{today}_{counter}.xlsx"
        if not candidate.exists():
            return candidate
        counter += 1


def export_summary_excel(rows: list[PaymentRow], folder_path: Path, total_sum: Decimal) -> Path:
    output_path = unique_output_path(folder_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Betalningar"

    sheet.append(OUTPUT_COLUMNS)
    # Löpnummer och Total visas bara på första raden per källfil. Raderna för
    # samma fil ligger i följd (stabil sortering, ett datum per fil), så ett
    # byte av source_file markerar början på en ny grupp.
    previous_source: str | None = None
    for row in rows:
        is_first_in_group = row.source_file != previous_source
        sheet.append(
            [
                row.datum,
                row.avsandare,
                row.betalningsreferens,
                float(row.belopp),
                row.lopnummer if is_first_in_group else None,
                (float(row.total) if row.total is not None else None) if is_first_in_group else None,
            ]
        )
        previous_source = row.source_file

    summary_row = sheet.max_row + 1
    sheet.cell(row=summary_row, column=1, value="SUMMA")
    sheet.cell(row=summary_row, column=6, value=float(total_sum))

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for cell in sheet[summary_row]:
        cell.font = Font(bold=True)

    for row_idx in range(2, summary_row):
        sheet.cell(row=row_idx, column=1).number_format = "yyyy-mm-dd"
        sheet.cell(row=row_idx, column=4).number_format = "#,##0.00"
        sheet.cell(row=row_idx, column=6).number_format = "#,##0.00"

    sheet.cell(row=summary_row, column=6).number_format = "#,##0.00"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:F{summary_row}"

    widths = {
        "A": 14,
        "B": 32,
        "C": 22,
        "D": 14,
        "E": 14,
        "F": 14,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    workbook.save(output_path)
    return output_path


def format_amount_sv(amount: Decimal) -> str:
    formatted = f"{amount:,.2f}"
    return formatted.replace(",", " ").replace(".", ",")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Sammanställ rensade betalningsfiler till en gemensam Excel-fil."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help=f'Mapp att läsa rensade filer från (matchar "{FILE_PATTERN}"). Standard: {BASE_DIR}.',
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Mapp att spara sammanställningen i. Standard: samma som --input.",
    )
    args = parser.parse_args()

    input_dir = args.input if args.input is not None else BASE_DIR
    output_dir = args.output if args.output is not None else input_dir

    if not input_dir.is_dir():
        print(f"FEL: Indatamappen finns inte: {input_dir}", file=sys.stderr)
        return 1
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:  # noqa: BLE001 - rapportera och avbryt snyggt.
        print(f"FEL: Kunde inte skapa utdatamappen: {exc}", file=sys.stderr)
        return 1

    files = find_matching_files(input_dir)
    if not files:
        print(f'Inga filer hittades som matchar "{FILE_PATTERN}" i {input_dir}.')
        return 1

    print(f"\nFiler som hittades ({len(files)}):")
    for file_path in files:
        print(f"  - {file_path.name}")

    results: list[FileResult] = []
    all_rows: list[PaymentRow] = []

    for file_path in files:
        print(f"\nBehandlar: {file_path.name}")
        result = collect_data_from_file(file_path)
        results.append(result)

        for warning_text in result.warnings:
            print(f"  VARNING: {warning_text}")
        for error_text in result.errors:
            print(f"  FEL: {error_text}")

        if not result.has_errors:
            print(f"  Betalningsrader: {len(result.rows)}")
            print(f"  Datum: {result.datum.isoformat() if result.datum else '(saknas)'}")
            print(f"  Löpnummer: {result.lopnummer or '(saknas)'}")
            print(f"  Total: {format_amount_sv(result.total) if result.total is not None else '(saknas)'}")
            all_rows.extend(result.rows)

    valid_rows = [row for row in all_rows if row.datum is not None and row.belopp is not None]
    if not valid_rows:
        print("\nInga giltiga betalningsrader hittades. Ingen output-fil skapades.")
        return 1

    valid_rows.sort(key=lambda row: row.datum, reverse=True)

    totals_by_file: dict[str, Decimal] = {}
    for result in results:
        if result.total is not None and result.rows and not result.has_errors:
            totals_by_file[result.path.name] = result.total
    total_sum = sum(totals_by_file.values(), Decimal("0.00")).quantize(Decimal("0.01"))

    output_path = export_summary_excel(valid_rows, output_dir, total_sum)

    print("\nSammanställning klar.")
    print(f"\nIndatamapp: {input_dir}")
    print(f"Filer hittade: {len(files)}")
    print(f"Filer behandlade: {sum(1 for result in results if not result.has_errors)}")
    print(f"Filer med varningar: {sum(1 for result in results if result.has_warnings)}")
    print(f"Filer med fel: {sum(1 for result in results if result.has_errors)}")
    print(f"Exporterade betalningsrader: {len(valid_rows)}")
    print(f"Total summa: {format_amount_sv(total_sum)}")
    print(f"Output-fil: {output_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
