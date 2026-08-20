#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
step_4_skapa_swish_csv.py

Skapar en CSV-fil för betalningsregistrering från samlade Swish-betalningar.

Följer samma princip som Bankgiro-steget (convert_betalningar_to_csv.py) och
producerar samma CSV-format, så att registreringssteget kan konsumera både
Bankgiro- och Swish-CSV likadant.

Indata-kolumner (Excel, samla_swish_bet_lista_*.xlsx):
    Bokförd | Avsändare | Meddelande | Insättningar | Period | Filtotal

Utdata-kolumner (CSV, exakt denna ordning):
    Datum;Avsändare;Betalningsreferens;Fakturanummer;Belopp

Mappning:
    Datum             <- Bokförd
    Avsändare         <- Avsändare
    Betalningsreferens<- Meddelande
    Fakturanummer     <- 5-siffrigt nummer extraherat ur Meddelande
    Belopp            <- Insättningar

Körs via dashboarden med --input/--output/--date, eller fristående:
    python step_4_skapa_swish_csv.py --input "C:\\...\\output"
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("FEL: Paketet 'openpyxl' saknas. Installera med:  pip install openpyxl")
    sys.exit(1)


# --- Konfiguration -----------------------------------------------------------

# Föredraget arknamn i den samlade Swish-filen. Saknas det används första arket.
SHEET_NAME = "Swish"

# Kolumner som MÅSTE finnas i Excel-filen (valideras vid inläsning).
REQUIRED_COLUMNS = ["Bokförd", "Avsändare", "Meddelande", "Insättningar"]

# Kolumner som ska skrivas till CSV, i exakt denna ordning (samma som Bankgiro).
CSV_COLUMNS = ["Datum", "Avsändare", "Betalningsreferens", "Fakturanummer", "Belopp"]

# När --input pekar på en mapp letar vi efter den samlade Swish-filen.
INPUT_PATTERN = "samla_swish_bet_lista_*.xlsx"
# Prefix för utdatafilerna.
CSV_PREFIX = "swish_betalningar_for_registrering"

# Fakturanummer är normalt exakt 5 siffror. Lookbehind/lookahead på siffror
# fångar "F49457" -> 49457 men matchar inte delar av längre tal.
INVOICE_RE = re.compile(r"(?<!\d)\d{5}(?!\d)")


# --- Hjälpfunktioner för normalisering ---------------------------------------

def normalize_cell_value(value: Any) -> str:
    """Gör om ett rått cellvärde till en ren textsträng.

    None -> "", riktiga datum -> YYYY-MM-DD, heltal utan ".0", text trimmas.
    """
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value:  # NaN
            return ""
        if value.is_integer():
            return str(int(value))
        return repr(value)
    return str(value).strip()


def format_amount(value: Any) -> str:
    """Formaterar Belopp för CSV utan att räkna om eller avrunda värdet."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value:  # NaN
            return ""
        if value.is_integer():
            return str(int(value))
        return repr(value)
    return str(value).strip()


def extract_invoice_number(reference: str) -> tuple[str, str, str]:
    """Extraherar ett fakturanummer (normalt 5 siffror) ur referensen.

    Returnerar (fakturanummer, status, meddelande), där status är "OK" eller
    "VARNING".
    """
    ref = (reference or "").strip()
    if not ref:
        return "", "VARNING", "Meddelande saknas – inget fakturanummer kunde extraheras"

    matches = INVOICE_RE.findall(ref)
    if not matches:
        return "", "VARNING", "Inget 5-siffrigt fakturanummer hittades i meddelandet"
    if len(matches) == 1:
        return matches[0], "OK", "Fakturanummer extraherat"

    chosen = matches[-1]
    msg = f"Flera 5-siffriga tal hittades ({', '.join(matches)}) – valde det sista ({chosen})"
    return chosen, "VARNING", msg


# --- Inläsning ---------------------------------------------------------------

def read_excel_file(path: str) -> tuple[list[dict], int]:
    """Läser Excel-filen och returnerar (rader, antal_lästa).

    Kolumner identifieras via namnen i rubrikraden, inte via fasta celladresser.
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Filen hittades inte: {path}")
    if not path.lower().endswith(".xlsx"):
        raise ValueError(f"Filen är inte en .xlsx-fil: {path}")

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"Kunde inte öppna Excel-filen: {exc}") from exc

    try:
        if not wb.sheetnames:
            raise ValueError("Excel-filen saknar ark.")

        sheet = SHEET_NAME if SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]
        rows_iter = ws.iter_rows(values_only=True)

        try:
            header = next(rows_iter)
        except StopIteration:
            raise ValueError("Excel-filen är tom (ingen rubrikrad hittades).")

        header_map: dict[str, int] = {}
        for idx, name in enumerate(header):
            if name is None:
                continue
            key = str(name).strip().lower()
            if key and key not in header_map:
                header_map[key] = idx

        missing = [c for c in REQUIRED_COLUMNS if c.strip().lower() not in header_map]
        if missing:
            raise ValueError(
                "Följande nödvändiga kolumner saknas i Excel-filen: " + ", ".join(missing)
            )

        col = {c: header_map[c.strip().lower()] for c in REQUIRED_COLUMNS}

        def cell(row: tuple, column_name: str) -> Any:
            i = col[column_name]
            return row[i] if i < len(row) else None

        records: list[dict] = []
        read_count = 0
        excel_row = 1  # rubriken låg på rad 1

        for row in rows_iter:
            excel_row += 1
            if row is None or all(v is None for v in row):
                continue

            read_count += 1

            bokford = cell(row, "Bokförd")
            avsandare = cell(row, "Avsändare")
            meddelande = cell(row, "Meddelande")
            insattningar = cell(row, "Insättningar")

            # En rad utan avsändare OCH meddelande betraktas som en icke-betalningsrad
            # (t.ex. SUMMA-raden, som bara har ett belopp) och exporteras inte.
            is_empty = (
                normalize_cell_value(avsandare) == ""
                and normalize_cell_value(meddelande) == ""
            )

            records.append(
                {
                    "excel_row": excel_row,
                    "Datum": bokford,
                    "Avsändare": avsandare,
                    "Betalningsreferens": meddelande,
                    "Belopp": insattningar,
                    "is_empty": is_empty,
                }
            )

        return records, read_count
    finally:
        wb.close()


def amount_to_decimal(value: Any) -> Decimal | None:
    """Tolka ett beloppsvärde till Decimal för summering (None om det inte går)."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value != value:  # NaN
            return None
        return Decimal(str(value))
    text = str(value).strip().replace(" ", "").replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def format_sum_sv(total: Decimal) -> str:
    """Formaterar en summa med svenskt format, t.ex. 41 899,00."""
    q = total.quantize(Decimal("0.01"))
    return f"{q:,.2f}".replace(",", " ").replace(".", ",")


# --- Bearbetning -------------------------------------------------------------

def process_records(records: list[dict]):
    """Normaliserar fält, extraherar fakturanummer, bygger CSV-rader och logg."""
    csv_rows: list[dict] = []
    log_entries: list[dict] = []
    found = 0
    missing = 0
    skipped = 0
    total_amount = Decimal("0.00")

    for rec in records:
        excel_row = rec["excel_row"]
        ref_text = normalize_cell_value(rec["Betalningsreferens"])

        if rec["is_empty"]:
            skipped += 1
            log_entries.append(
                {
                    "excel_row": excel_row,
                    "referens": ref_text,
                    "fakturanummer": "",
                    "status": "ÖVERHOPPAD",
                    "meddelande": "Raden saknar avsändare och meddelande – exporteras inte (t.ex. summarad)",
                }
            )
            continue

        datum = normalize_cell_value(rec["Datum"])
        avsandare = normalize_cell_value(rec["Avsändare"])
        belopp = format_amount(rec["Belopp"])

        belopp_dec = amount_to_decimal(rec["Belopp"])
        if belopp_dec is not None:
            total_amount += belopp_dec

        fakturanummer, status, meddelande = extract_invoice_number(ref_text)
        if fakturanummer:
            found += 1
        else:
            missing += 1

        csv_rows.append(
            {
                "Datum": datum,
                "Avsändare": avsandare,
                "Betalningsreferens": ref_text,
                "Fakturanummer": fakturanummer,
                "Belopp": belopp,
            }
        )
        log_entries.append(
            {
                "excel_row": excel_row,
                "referens": ref_text,
                "fakturanummer": fakturanummer,
                "status": status,
                "meddelande": meddelande,
            }
        )

    return csv_rows, log_entries, found, missing, skipped, total_amount


# --- Utdata ------------------------------------------------------------------

def export_csv(csv_rows: list[dict], out_path: str) -> None:
    """Skriver CSV-filen: semikolon, UTF-8 med BOM, Windows-radslut (CRLF)."""
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        writer.writerow(CSV_COLUMNS)
        for row in csv_rows:
            writer.writerow([row[c] for c in CSV_COLUMNS])


def write_log(log_entries: list[dict], summary: dict, log_path: str) -> None:
    """Skriver en loggfil med en rad per Excel-rad samt en sammanfattning."""
    with open(log_path, "w", encoding="utf-8-sig", newline="") as f:
        f.write("Loggfil – konvertering av Swish-betalningar till CSV\n")
        f.write(f"Skapad:  {summary['timestamp_human']}\n")
        f.write(f"Källfil: {summary['excel_path']}\n")
        f.write(f"CSV-fil: {summary['csv_path']}\n")
        f.write("\n")
        f.write("Sammanfattning\n")
        f.write(f"  Rader lästa:           {summary['read']}\n")
        f.write(f"  Rader exporterade:     {summary['exported']}\n")
        f.write(f"  Fakturanummer hittade: {summary['found']}\n")
        f.write(f"  Fakturanummer saknas:  {summary['missing']}\n")
        f.write(f"  Rader överhoppade:     {summary['skipped']}\n")
        f.write(f"  Summa belopp:          {summary['total_amount']}\n")
        f.write("\n")
        f.write("Detaljer per rad\n")
        f.write(f"{'Excel-rad':>9} | {'Status':<11} | {'Fakturanr':<10} | Betalningsreferens\n")
        f.write("-" * 90 + "\n")
        for e in log_entries:
            f.write(f"{e['excel_row']:>9} | {e['status']:<11} | {e['fakturanummer']:<10} | {e['referens']}\n")
            if e["status"] != "OK":
                f.write(f"{'':>9} | {'':<11} | {'':<10} | -> {e['meddelande']}\n")


# --- Indata / utdata-vägar ---------------------------------------------------

def resolve_excel_path(raw: str) -> str:
    """Tolka --input till en konkret .xlsx-fil. En mapp -> senaste INPUT_PATTERN."""
    cleaned = raw.strip().strip('"').strip("'").strip()
    path = Path(cleaned)
    if path.is_dir():
        candidates = [c for c in path.glob(INPUT_PATTERN)
                      if c.is_file() and not c.name.startswith("~$")]
        if not candidates:
            raise FileNotFoundError(f'Inga filer matchade "{INPUT_PATTERN}" i mappen: {path}')
        return str(max(candidates, key=lambda c: c.stat().st_mtime))
    return str(path)


def build_output_paths(out_dir: Path, date_str: str) -> tuple[Path, Path]:
    """Bygg unika sökvägar för CSV och loggfil utifrån datum."""
    csv_path = out_dir / f"{CSV_PREFIX}_{date_str}.csv"
    counter = 2
    while csv_path.exists():
        csv_path = out_dir / f"{CSV_PREFIX}_{date_str}_{counter}.csv"
        counter += 1
    log_path = out_dir / f"{csv_path.stem}_log.txt"
    return csv_path, log_path


# --- Huvudprogram ------------------------------------------------------------

def main() -> int:
    print("=" * 60)
    print(" Steg 4: Samlade Swish-betalningar (.xlsx) -> CSV")
    print("=" * 60)

    parser = argparse.ArgumentParser(
        description="Skapa CSV för betalningsregistrering från samlade Swish-betalningar."
    )
    parser.add_argument(
        "--input",
        type=str,
        default=None,
        help=f'Mapp eller .xlsx-fil. En mapp läser senast ändrade "{INPUT_PATTERN}".',
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Mapp att spara CSV + logg i. Standard: samma mapp som indatafilen.",
    )
    parser.add_argument(
        "--date",
        type=str,
        default=None,
        help="Datum (YYYY-MM-DD) som används i CSV-filnamnet. Standard: dagens datum.",
    )
    parser.add_argument(
        "path",
        nargs="?",
        default=None,
        help="Valfri direkt sökväg till en .xlsx-fil (bakåtkompatibelt).",
    )
    args = parser.parse_args()

    raw_input = args.input if args.input else args.path
    if not raw_input:
        print("FEL: Ange indata med --input <mapp eller .xlsx-fil>.", file=sys.stderr)
        return 1

    try:
        excel_path = resolve_excel_path(raw_input)
    except FileNotFoundError as exc:
        print(f"FEL: {exc}", file=sys.stderr)
        return 1

    print(f"Indatafil: {excel_path}")

    try:
        records, read_count = read_excel_file(excel_path)
    except FileNotFoundError as exc:
        print(f"FEL: {exc}", file=sys.stderr)
        return 1
    except ValueError as exc:
        print(f"FEL: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:  # oväntat fel – visa tydligt utan ful stacktrace
        print(f"OVÄNTAT FEL vid inläsning: {exc}", file=sys.stderr)
        return 1

    csv_rows, log_entries, found, missing, skipped, total_amount = process_records(records)
    exported = len(csv_rows)
    total_amount_str = format_sum_sv(total_amount)

    date_str = (args.date or "").strip() or dt.date.today().strftime("%Y-%m-%d")

    if args.output:
        out_dir = Path(args.output.strip().strip('"').strip("'").strip())
    else:
        out_dir = Path(excel_path).resolve().parent
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:  # noqa: BLE001
        print(f"FEL: Kunde inte skapa utdatamappen: {exc}", file=sys.stderr)
        return 1

    csv_path, log_path = build_output_paths(out_dir, date_str)

    now = dt.datetime.now()
    summary = {
        "timestamp_human": now.strftime("%Y-%m-%d %H:%M:%S"),
        "excel_path": os.path.abspath(excel_path),
        "csv_path": str(csv_path),
        "read": read_count,
        "exported": exported,
        "found": found,
        "missing": missing,
        "skipped": skipped,
        "total_amount": total_amount_str,
    }

    try:
        export_csv(csv_rows, str(csv_path))
        write_log(log_entries, summary, str(log_path))
    except Exception as exc:
        print(f"FEL vid skrivning av utdatafiler: {exc}", file=sys.stderr)
        return 1

    print("\nKlart!")
    print(f"  Rader lästa:           {read_count}")
    print(f"  Rader exporterade:     {exported}")
    print(f"  Fakturanummer hittade: {found}")
    print(f"  Fakturanummer saknas:  {missing}")
    if skipped:
        print(f"  Rader överhoppade:     {skipped}  (tom-/summarad)")
    print(f"  Summa belopp:          {total_amount_str}")
    print(f"\n  CSV:  {csv_path}")
    print(f"  Logg: {log_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
