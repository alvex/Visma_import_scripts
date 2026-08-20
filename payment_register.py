from __future__ import annotations

import csv
import random
import re
import sys
import time
from dataclasses import asdict, dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from playwright.sync_api import (
    BrowserContext,
    Error as PlaywrightError,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)


BASE_URL = "https://hemfresh.com/fakturorab/index.php"
PAYMENT_URL_TEMPLATE = (
    "https://hemfresh.com/fakturorab/index.php"
    "?module=payments&view=process&invoice={invoice}&op=pay_selected_invoice"
)

SCRIPT_DIR = Path(__file__).resolve().parent
AUTH_STATE_PATH = SCRIPT_DIR / "auth_state.json"
SCREENSHOT_DIR = SCRIPT_DIR / "screenshots"
EXCEL_PATTERN = "total_betalningar_*.xlsx"

REQUIRED_COLUMNS = ["Datum", "Avsändare", "Betalningsreferens", "Belopp"]
DEFAULT_PAYMENT_METHOD = "Bankbetalning"
DEFAULT_NOTE_PREFIX = "Automatisk betalningsregistrering"


@dataclass
class PaymentRow:
    row_number: int
    datum: date
    avsandare: str
    betalningsreferens: str
    belopp: Decimal
    fakturanummer: str


@dataclass
class PaymentLogRow:
    Datum: str
    Avsändare: str
    Betalningsreferens: str
    Fakturanummer: str
    Belopp_fran_Excel: str
    Belopp_pa_fakturasidan: str
    Status: str
    Meddelande: str
    Tidpunkt: str
    Screenshot: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_key(value: Any) -> str:
    return normalize_text(value).casefold()


def normalize_amount(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, int | float):
        return Decimal(str(value)).quantize(Decimal("0.01"))

    text = normalize_text(value)
    text = re.sub(r"(?i)\b(sek|kr|kronor)\b", "", text)
    text = re.sub(r"[^0-9,.\- ]", "", text).replace(" ", "")
    if not text or text in {"-", ".", ","}:
        return None

    if "," in text and "." in text:
        if text.rfind(".") > text.rfind(","):
            text = text.replace(",", "")
        else:
            text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        if re.fullmatch(r"-?\d{1,3}(,\d{3})+", text):
            text = text.replace(",", "")
        else:
            text = text.replace(".", "").replace(",", ".")
    elif text.count(".") > 1:
        parts = text.split(".")
        text = "".join(parts[:-1]) + "." + parts[-1]

    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def normalize_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None

    text = normalize_text(value)
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def amount_to_text(amount: Decimal | None) -> str:
    return "" if amount is None else f"{amount:.2f}"


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def ask_for_folder_path() -> Path | None:
    raw = input("Ange sökvägen till mappen där total_betalningar_*.xlsx finns: ").strip()
    raw = raw.strip('"').strip("'")
    folder_path = Path(raw)

    if not folder_path.exists() or not folder_path.is_dir():
        print(f"Fel: Mappen finns inte: {folder_path}", file=sys.stderr)
        return None
    return folder_path


def find_latest_excel_file(folder_path: Path) -> Path | None:
    files = [
        path
        for path in folder_path.glob(EXCEL_PATTERN)
        if path.is_file() and not path.name.startswith("~$")
    ]

    if not files:
        print(f"Fel: Ingen fil hittades som matchar {EXCEL_PATTERN} i {folder_path}.", file=sys.stderr)
        return None

    latest_file = max(files, key=lambda path: path.stat().st_mtime)
    if len(files) > 1:
        print(f"Flera filer hittades. Senast ändrade fil väljs: {latest_file}")
    else:
        print(f"Excel-fil som används: {latest_file}")
    return latest_file


def ask_for_run_mode() -> bool:
    print("\nVälj körläge:")
    print("1. Dry-run/testläge")
    print("2. Riktig registrering")
    choice = input("Ange 1 eller 2: ").strip() or "1"

    if choice != "2":
        return True

    confirmation = input('Skriv REGISTRERA för att bekräfta riktig registrering: ').strip()
    if confirmation == "REGISTRERA":
        return False

    print("Bekräftelsen matchade inte. Körningen fortsätter i dry-run/testläge.")
    return True


def ask_for_row_limit() -> int | None:
    raw = input("\nHur många rader vill du behandla? Skriv 1 för första testet eller ALLA för alla rader: ").strip()
    if not raw:
        return 1
    if raw.casefold() == "alla":
        return None

    try:
        value = int(raw)
        return max(value, 1)
    except ValueError:
        print("Ogiltigt val. Standard används: 1 rad.")
        return 1


def extract_invoice_number(reference: str) -> str | None:
    text = normalize_text(reference)
    if not text:
        return None

    patterns = [
        r"(?i)\bfaktura\s*(?:nr|nummer|#|:)?\s*(\d{4,7})\b",
        r"(?i)\bfak(?:tura)?(?:nr|t\.?|\.?)?\s*(?:nr|nummer|#|:)?\s*(\d{4,7})\b",
        r"(?i)\bocr\s*(?:nr|nummer|#|:)?\s*(\d{4,7})\b",
        r"(?i)\binvoice\s*(?:no|nr|number|#|:)?\s*(\d{4,7})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)

    candidates = re.findall(r"(?<!\d)(\d{4,7})(?!\d)", text)
    candidates = [candidate for candidate in candidates if not candidate.startswith(("19", "20"))]
    unique_candidates = sorted(set(candidates))
    if len(unique_candidates) == 1:
        return unique_candidates[0]
    return None


def validate_columns(column_map: dict[str, int]) -> None:
    missing = [name for name in REQUIRED_COLUMNS if name not in column_map]
    if missing:
        raise ValueError(f"Excel-filen saknar kolumner: {', '.join(missing)}")


def read_excel_file(excel_path: Path) -> tuple[list[PaymentRow], list[PaymentLogRow]]:
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active

    header_row: int | None = None
    column_map: dict[str, int] = {}
    required_by_key = {normalize_key(name): name for name in REQUIRED_COLUMNS}

    for row_idx in range(1, sheet.max_row + 1):
        found: dict[str, int] = {}
        for cell in sheet[row_idx]:
            key = normalize_key(cell.value)
            if key in required_by_key:
                found[required_by_key[key]] = cell.column
        if all(name in found for name in REQUIRED_COLUMNS):
            header_row = row_idx
            column_map = found
            break

    validate_columns(column_map)
    assert header_row is not None

    payments: list[PaymentRow] = []
    logs: list[PaymentLogRow] = []

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        values = {name: sheet.cell(row=row_idx, column=column_map[name]).value for name in REQUIRED_COLUMNS}
        if all(value in (None, "") for value in values.values()):
            continue

        if normalize_key(values["Datum"]) == "summa":
            continue

        datum = normalize_date(values["Datum"])
        amount = normalize_amount(values["Belopp"])
        avsandare = normalize_text(values["Avsändare"])
        reference = normalize_text(values["Betalningsreferens"])
        invoice = extract_invoice_number(reference)

        if datum is None:
            logs.append(make_log(values["Datum"], avsandare, reference, invoice or "", amount, None, "SAKNAR_DATUM", "Datum saknas eller kan inte tolkas."))
            continue
        if amount is None:
            logs.append(make_log(datum.isoformat(), avsandare, reference, invoice or "", amount, None, "FEL", "Belopp saknas eller kan inte tolkas."))
            continue
        if invoice is None:
            logs.append(make_log(datum.isoformat(), avsandare, reference, "", amount, None, "SAKNAR_FAKTURANUMMER", "Fakturanummer kunde inte tolkas säkert."))
            continue

        payments.append(
            PaymentRow(
                row_number=row_idx,
                datum=datum,
                avsandare=avsandare,
                betalningsreferens=reference,
                belopp=amount,
                fakturanummer=invoice,
            )
        )

    return payments, logs


def make_log(
    datum: Any,
    avsandare: str,
    reference: str,
    invoice: str,
    excel_amount: Decimal | None,
    page_amount: Decimal | None,
    status: str,
    message: str,
    screenshot: str = "",
) -> PaymentLogRow:
    return PaymentLogRow(
        Datum=normalize_text(datum),
        Avsändare=avsandare,
        Betalningsreferens=reference,
        Fakturanummer=invoice,
        Belopp_fran_Excel=amount_to_text(excel_amount),
        Belopp_pa_fakturasidan=amount_to_text(page_amount),
        Status=status,
        Meddelande=message,
        Tidpunkt=now_text(),
        Screenshot=screenshot,
    )


def load_or_create_browser_session(playwright) -> tuple[BrowserContext, Page]:
    browser = playwright.chromium.launch(headless=False)
    context_kwargs = {"storage_state": str(AUTH_STATE_PATH)} if AUTH_STATE_PATH.exists() else {}
    context = browser.new_context(**context_kwargs)
    page = context.new_page()

    page.goto(BASE_URL, wait_until="domcontentloaded", timeout=30000)
    print("\nWebbläsaren är öppen.")
    print("Om du inte är inloggad: logga in manuellt i fakturasystemet.")
    input("Tryck Enter här när fakturasystemet visas och du är inloggad...")
    context.storage_state(path=str(AUTH_STATE_PATH))
    print(f"Session sparad/uppdaterad: {AUTH_STATE_PATH}")

    return context, page


def open_payment_page(page: Page, invoice: str) -> None:
    page.goto(PAYMENT_URL_TEMPLATE.format(invoice=invoice), wait_until="domcontentloaded", timeout=30000)


def page_text(page: Page) -> str:
    return page.locator("body").inner_text(timeout=10000)


def read_amount_to_pay(page: Page) -> Decimal | None:
    text = page_text(page)
    match = re.search(r"Att\s+betala\s*:?\s*([\-0-9\s.,]+(?:kr|SEK)?)", text, flags=re.IGNORECASE)
    if not match:
        return None
    return normalize_amount(match.group(1))


def read_invoice_number_from_page(page: Page) -> str | None:
    text = page_text(page)
    patterns = [
        r"ID\s+Faktura\s*:?\s*(\d{4,7})",
        r"Faktura(?:nr)?\s*:?\s*(\d{4,7})",
        r"Invoice\s*:?\s*(\d{4,7})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return match.group(1)
    return None


def page_looks_paid(page: Page, amount_to_pay: Decimal | None) -> bool:
    text = page_text(page).casefold()
    paid_markers = ["redan betald", "är betald", "betald faktura", "skuld: 0.00", "att betala: 0"]
    return amount_to_pay == Decimal("0.00") or any(marker in text for marker in paid_markers)


def locator_exists(page: Page, selector: str) -> bool:
    try:
        locator = page.locator(selector).first
        return locator.count() > 0 and locator.is_visible(timeout=1000)
    except Exception:
        return False


def find_first_visible(page: Page, selectors: list[str]):
    for selector in selectors:
        try:
            locator = page.locator(selector).first
            if locator.count() > 0 and locator.is_visible(timeout=1000):
                return locator
        except Exception:
            continue
    return None


def check_invoice_page(page: Page, expected_invoice: str) -> tuple[bool, str, Decimal | None]:
    try:
        amount_to_pay = read_amount_to_pay(page)
    except PlaywrightTimeoutError:
        return False, "TIMEOUT", None

    if amount_to_pay is None:
        return False, "SAKNAR_BELOPP", None

    page_invoice = read_invoice_number_from_page(page)
    if page_invoice is not None and page_invoice != expected_invoice:
        return False, "FEL_FAKTURA", amount_to_pay

    if page_looks_paid(page, amount_to_pay):
        return False, "REDAN_BETALD", amount_to_pay

    amount_input = find_first_visible(page, amount_selectors())
    date_input = find_first_visible(page, date_selectors())
    submit_button = find_first_visible(page, submit_selectors())

    if amount_input is None or date_input is None or submit_button is None:
        return False, "SAKNAR_FORMULAR", amount_to_pay

    return True, "OK", amount_to_pay


def amount_selectors() -> list[str]:
    return [
        'input[name="amount"]',
        'input[id="amount"]',
        'input[name*="amount" i]',
        'input[id*="amount" i]',
        'input[name*="belopp" i]',
        'input[id*="belopp" i]',
    ]


def date_selectors() -> list[str]:
    return [
        'input[type="date"]',
        'input[name="date"]',
        'input[id="date"]',
        'input[name*="date" i]',
        'input[id*="date" i]',
        'input[name*="datum" i]',
        'input[id*="datum" i]',
    ]


def payment_method_selectors() -> list[str]:
    return [
        'select[name*="method" i]',
        'select[id*="method" i]',
        'select[name*="metodo" i]',
        'select[id*="metodo" i]',
        'select[name*="type" i]',
        'select[id*="type" i]',
    ]


def note_selectors() -> list[str]:
    return [
        'textarea[name*="note" i]',
        'textarea[id*="note" i]',
        'textarea[name*="anteckning" i]',
        'textarea[id*="anteckning" i]',
        'input[name*="note" i]',
        'input[id*="note" i]',
    ]


def submit_selectors() -> list[str]:
    return [
        'button:has-text("Betalningsprocess")',
        'input[type="submit"][value*="Betalningsprocess" i]',
        'button:has-text("Registrera")',
        'button:has-text("Spara")',
        'input[type="submit"][value*="Registrera" i]',
        'input[type="submit"][value*="Spara" i]',
    ]


def fill_payment_form(page: Page, payment: PaymentRow) -> tuple[bool, str]:
    amount_input = find_first_visible(page, amount_selectors())
    date_input = find_first_visible(page, date_selectors())
    if amount_input is None or date_input is None:
        return False, "Saknar fält för Belopp eller Datum."

    amount_input.fill(f"{payment.belopp:.2f}")
    date_input.fill(payment.datum.isoformat())

    method_select = find_first_visible(page, payment_method_selectors())
    if method_select is not None:
        try:
            method_select.select_option(label=DEFAULT_PAYMENT_METHOD)
        except PlaywrightError:
            try:
                method_select.select_option(value=DEFAULT_PAYMENT_METHOD)
            except PlaywrightError:
                return False, f"Kunde inte välja betalningsmetod {DEFAULT_PAYMENT_METHOD}."

    note_input = find_first_visible(page, note_selectors())
    if note_input is not None:
        note = (
            f"{DEFAULT_NOTE_PREFIX}. Avsändare: {payment.avsandare}. "
            f"Referens: {payment.betalningsreferens}. Importerad: {now_text()}."
        )
        note_input.fill(note)

    return True, "Formuläret fylldes i."


def submit_payment(page: Page) -> bool:
    submit_button = find_first_visible(page, submit_selectors())
    if submit_button is None:
        return False
    submit_button.click()
    return True


def verify_payment_result(page: Page) -> bool:
    try:
        page.wait_for_load_state("domcontentloaded", timeout=15000)
        text = page_text(page).casefold()
    except Exception:
        return False

    success_markers = ["registrerad", "sparad", "betalningen", "redan betald", "betald"]
    return any(marker in text for marker in success_markers)


def take_error_screenshot(page: Page, invoice: str, status: str) -> str:
    SCREENSHOT_DIR.mkdir(exist_ok=True)
    safe_invoice = invoice or "utan_fakturanummer"
    path = SCREENSHOT_DIR / f"{safe_invoice}_{status}_{datetime.now():%Y-%m-%d_%H%M%S}.png"
    try:
        page.screenshot(path=str(path), full_page=True)
        return str(path)
    except Exception:
        return ""


def process_payment(page: Page, payment: PaymentRow, dry_run: bool) -> PaymentLogRow:
    page_amount: Decimal | None = None
    status = "FEL"
    screenshot = ""

    try:
        open_payment_page(page, payment.fakturanummer)
        page_ok, page_status, page_amount = check_invoice_page(page, payment.fakturanummer)

        if not page_ok:
            if page_status not in {"REDAN_BETALD"}:
                screenshot = take_error_screenshot(page, payment.fakturanummer, page_status)
            return make_log(
                payment.datum.isoformat(),
                payment.avsandare,
                payment.betalningsreferens,
                payment.fakturanummer,
                payment.belopp,
                page_amount,
                page_status,
                f"Sidan klarade inte säkerhetskontrollen: {page_status}.",
                screenshot,
            )

        if page_amount != payment.belopp:
            status = "BELOPP_MATCHAR_INTE"
            screenshot = take_error_screenshot(page, payment.fakturanummer, status)
            return make_log(
                payment.datum.isoformat(),
                payment.avsandare,
                payment.betalningsreferens,
                payment.fakturanummer,
                payment.belopp,
                page_amount,
                status,
                "Belopp från Excel matchar inte Att betala på fakturasidan.",
                screenshot,
            )

        form_ok, form_message = fill_payment_form(page, payment)
        if not form_ok:
            status = "SAKNAR_FORMULAR"
            screenshot = take_error_screenshot(page, payment.fakturanummer, status)
            return make_log(
                payment.datum.isoformat(),
                payment.avsandare,
                payment.betalningsreferens,
                payment.fakturanummer,
                payment.belopp,
                page_amount,
                status,
                form_message,
                screenshot,
            )

        if dry_run:
            message = (
                f"Skulle registrera betalning för faktura {payment.fakturanummer} "
                f"med belopp {payment.belopp:.2f} och datum {payment.datum.isoformat()}."
            )
            return make_log(
                payment.datum.isoformat(),
                payment.avsandare,
                payment.betalningsreferens,
                payment.fakturanummer,
                payment.belopp,
                page_amount,
                "DRY_RUN_OK",
                message,
            )

        if not submit_payment(page):
            status = "SAKNAR_FORMULAR"
            screenshot = take_error_screenshot(page, payment.fakturanummer, status)
            return make_log(
                payment.datum.isoformat(),
                payment.avsandare,
                payment.betalningsreferens,
                payment.fakturanummer,
                payment.belopp,
                page_amount,
                status,
                "Knappen Betalningsprocess hittades inte.",
                screenshot,
            )

        if verify_payment_result(page):
            status = "REGISTRERAD"
            message = "Betalningen verkar vara registrerad."
        else:
            status = "OKLAR_STATUS"
            message = "Formuläret skickades, men resultatet kunde inte verifieras säkert."

        return make_log(
            payment.datum.isoformat(),
            payment.avsandare,
            payment.betalningsreferens,
            payment.fakturanummer,
            payment.belopp,
            page_amount,
            status,
            message,
        )

    except PlaywrightTimeoutError as exc:
        status = "TIMEOUT"
        screenshot = take_error_screenshot(page, payment.fakturanummer, status)
        return make_log(payment.datum.isoformat(), payment.avsandare, payment.betalningsreferens, payment.fakturanummer, payment.belopp, page_amount, status, str(exc), screenshot)
    except Exception as exc:  # noqa: BLE001 - one failed row must not stop the batch.
        status = "FEL"
        screenshot = take_error_screenshot(page, payment.fakturanummer, status)
        return make_log(payment.datum.isoformat(), payment.avsandare, payment.betalningsreferens, payment.fakturanummer, payment.belopp, page_amount, status, str(exc), screenshot)


def save_log(logs: list[PaymentLogRow], folder_path: Path) -> Path:
    log_path = folder_path / f"payment_log_{datetime.now():%Y-%m-%d_%H%M%S}.csv"
    fieldnames = list(asdict(logs[0]).keys())
    with log_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        for row in logs:
            writer.writerow(asdict(row))
    return log_path


def main() -> int:
    folder_path = ask_for_folder_path()
    if folder_path is None:
        return 1

    excel_path = find_latest_excel_file(folder_path)
    if excel_path is None:
        return 1

    dry_run = ask_for_run_mode()
    row_limit = ask_for_row_limit()

    try:
        payments, logs = read_excel_file(excel_path)
    except Exception as exc:
        print(f"Fel vid läsning av Excel-filen: {exc}", file=sys.stderr)
        return 1

    if row_limit is not None:
        payments_to_process = payments[:row_limit]
    else:
        payments_to_process = payments

    print(f"\nGiltiga betalningsrader: {len(payments)}")
    print(f"Rader som behandlas: {len(payments_to_process)}")
    print(f"Körläge: {'Dry-run/testläge' if dry_run else 'Riktig registrering'}")

    if not payments_to_process:
        print("Inga giltiga betalningsrader att behandla.")
        if logs:
            log_path = save_log(logs, folder_path)
            print(f"Loggfil: {log_path}")
        return 1

    with sync_playwright() as playwright:
        context, page = load_or_create_browser_session(playwright)

        for index, payment in enumerate(payments_to_process, start=1):
            print(f"\n[{index}/{len(payments_to_process)}] Faktura {payment.fakturanummer}, belopp {payment.belopp:.2f}")
            log_row = process_payment(page, payment, dry_run)
            logs.append(log_row)
            print(f"  {log_row.Status}: {log_row.Meddelande}")
            time.sleep(random.uniform(1.0, 2.0))

        browser = context.browser
        context.storage_state(path=str(AUTH_STATE_PATH))
        context.close()
        if browser is not None:
            browser.close()

    log_path = save_log(logs, folder_path)
    print("\nKörning klar.")
    print(f"Loggfil: {log_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
