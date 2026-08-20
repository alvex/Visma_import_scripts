"""Rensa Swish-transaktionsfiler till en standardiserad lista.

Läser filer som matchar "Transaktioner * Swish*.xlsx", behåller endast
kolumnerna Bokförd, Typ, Avsändare, Meddelande, Insättningar och Summa, och
skapar en ny fil clean_swish_lista_<periodslut>.xlsx där perioden står överst,
före själva tabellen.

Körs via dashboarden med --input/--output, eller fristående från terminalen:
    python clean_swish_files.py --input "C:\\...\\data" --output "C:\\...\\edit"
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

BASE_DIR = Path(__file__).resolve().parent
FILE_PATTERN = "Transaktioner * Swish*.xlsx"
OUTPUT_PREFIX = "clean_swish_lista_"

# Kolumner som ska behållas i resultatfilen, i denna ordning. Alla andra tas bort.
WANTED_COLUMNS = ["Bokförd", "Typ", "Avsändare", "Meddelande", "Insättningar", "Summa"]
AMOUNT_COLUMNS = {"Insättningar", "Summa"}
DATE_COLUMNS = {"Bokförd"}

PERIOD_RE = re.compile(r"period:\s*(\d{4}-\d{2}-\d{2})\s*-\s*(\d{4}-\d{2}-\d{2})", re.IGNORECASE)
ISODATE_RE = re.compile(r"\d{4}-\d{2}-\d{2}")


# ---------------------------------------------------------------------------
# Hjälpfunktioner för rensning
# ---------------------------------------------------------------------------
def normalize_text(value: Any) -> str:
    """Trimma och kollapsa extra mellanslag. None -> ''."""
    if value is None:
        return ""
    text = str(value).replace(" ", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def header_key(value: Any) -> str:
    return normalize_text(value).casefold()


def format_date(value: Any) -> str | None:
    """Returnera datum i tydligt format YYYY-MM-DD."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = normalize_text(value)
    if not text:
        return None
    match = ISODATE_RE.search(text)
    return match.group(0) if match else text


def clean_amount(value: Any):
    """Behåll tal som tal (formateras med 2 decimaler), och svenska
    kommasträngar (t.ex. "547,00") som de är."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = normalize_text(value)
    return text or None


def clean_value(column_name: str, value: Any):
    if column_name in DATE_COLUMNS:
        return format_date(value)
    if column_name in AMOUNT_COLUMNS:
        return clean_amount(value)
    return normalize_text(value) or None


# ---------------------------------------------------------------------------
# Hitta period och tabell
# ---------------------------------------------------------------------------
def find_period(sheet: Worksheet) -> tuple[str | None, str | None, str | None]:
    """Returnera (period_text, startdatum, slutdatum). Saknas perioden -> None."""
    for row in sheet.iter_rows():
        for cell in row:
            text = normalize_text(cell.value)
            if text.lower().startswith("period:"):
                match = PERIOD_RE.search(text)
                if match:
                    return text, match.group(1), match.group(2)
                return text, None, None
    return None, None, None


def find_header_row(sheet: Worksheet) -> tuple[int | None, dict[str, int], int]:
    """Hitta rubrikraden (den rad som innehåller flest av de önskade kolumnerna).
    Returnerar (radnummer, {kolumnnamn: kolumnindex}, antal_träffar)."""
    wanted = {header_key(name): name for name in WANTED_COLUMNS}
    best_row: int | None = None
    best_map: dict[str, int] = {}
    best_count = 0
    for row_idx in range(1, sheet.max_row + 1):
        found: dict[str, int] = {}
        for cell in sheet[row_idx]:
            key = header_key(cell.value)
            if key in wanted:
                found[wanted[key]] = cell.column
        if len(found) > best_count:
            best_count = len(found)
            best_row = row_idx
            best_map = found
    return best_row, best_map, best_count


def is_duplicate_header(values: dict[str, Any]) -> bool:
    """True om raden bara upprepar kolumnrubrikerna."""
    return all(header_key(values[name]) == header_key(name) for name in WANTED_COLUMNS)


# ---------------------------------------------------------------------------
# Skriv resultatfil
# ---------------------------------------------------------------------------
def write_output(period_line: str, rows: list[dict[str, Any]], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Swish"

    # Perioden överst, sedan en tom rad, sedan tabellen.
    sheet.cell(row=1, column=1, value=period_line).font = Font(bold=True)
    for col_idx, name in enumerate(WANTED_COLUMNS, start=1):
        cell = sheet.cell(row=3, column=col_idx, value=name)
        cell.font = Font(bold=True)

    row_idx = 4
    for values in rows:
        for col_idx, name in enumerate(WANTED_COLUMNS, start=1):
            value = values[name]
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if name in AMOUNT_COLUMNS and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
        row_idx += 1

    widths = {"A": 12, "B": 16, "C": 30, "D": 42, "E": 14, "F": 14}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def unique_path(output_dir: Path, end_date: str) -> Path:
    base = output_dir / f"{OUTPUT_PREFIX}{end_date}.xlsx"
    if not base.exists():
        return base
    counter = 2
    while True:
        candidate = output_dir / f"{OUTPUT_PREFIX}{end_date}_{counter}.xlsx"
        if not candidate.exists():
            return candidate
        counter += 1


# ---------------------------------------------------------------------------
# Bearbeta en fil
# ---------------------------------------------------------------------------
def process_file(src_path: Path, output_dir: Path) -> bool:
    """Returnerar True om en resultatfil skapades."""
    print(f"Hittade fil: {src_path.name}")

    try:
        workbook = load_workbook(src_path, data_only=True)
    except Exception as exc:  # noqa: BLE001 - rapportera och fortsätt med nästa fil.
        print(f"Fel: Kunde inte öppna filen: {exc}", file=sys.stderr)
        return False

    sheet = workbook.worksheets[0]

    period_text, _period_start, period_end = find_period(sheet)
    if period_text:
        print(f"Period hittad: {period_text.split(':', 1)[1].strip()}")
        period_line = period_text
        end_date = period_end or date.today().isoformat()
        if period_end is None:
            print("Fel: Perioden kunde inte tolkas, använder dagens datum i filnamnet.", file=sys.stderr)
    else:
        print("Fel: Ingen period hittades, använder dagens datum i filnamnet.", file=sys.stderr)
        end_date = date.today().isoformat()
        period_line = f"Period: {end_date}"

    header_row, col_map, count = find_header_row(sheet)
    if header_row is None or count < 2:
        print("Fel: Kunde inte hitta tabellen med Swish-transaktioner i filen.", file=sys.stderr)
        return False

    missing = [name for name in WANTED_COLUMNS if name not in col_map]
    if missing:
        for name in missing:
            print(f'Fel: Kolumnen "{name}" saknas.', file=sys.stderr)
        return False

    rows: list[dict[str, Any]] = []
    transaction_count = 0
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        values = {name: clean_value(name, sheet.cell(row=row_idx, column=col_map[name]).value)
                  for name in WANTED_COLUMNS}
        if all(value in (None, "") for value in values.values()):
            continue  # ta bort helt tomma rader
        if is_duplicate_header(values):
            continue  # ta bort dubbla rubrikrader
        rows.append(values)
        if values["Bokförd"] or values["Avsändare"]:
            transaction_count += 1

    if not rows:
        print("Fel: Inga transaktionsrader hittades under rubrikraden.", file=sys.stderr)
        return False

    print(f"Antal transaktioner: {transaction_count}")

    output_path = unique_path(output_dir, end_date)
    write_output(period_line, rows, output_path)
    print(f"Skapad fil: {output_path}")
    return True


def find_matching_files(input_dir: Path) -> list[Path]:
    return sorted(
        (path for path in input_dir.glob(FILE_PATTERN)
         if path.is_file() and not path.name.startswith("~$")),
        key=lambda path: path.name.casefold(),
    )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Rensa Swish-transaktionsfiler till clean_swish_lista_<periodslut>.xlsx."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help=f'Mapp med råfiler som matchar "{FILE_PATTERN}". Standard: {BASE_DIR}.',
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Mapp att spara rensade filer i. Standard: <input>/edit.",
    )
    args = parser.parse_args()

    input_dir = args.input if args.input is not None else BASE_DIR
    output_dir = args.output if args.output is not None else (input_dir / "edit")

    if not input_dir.is_dir():
        print(f"Fel: Indatamappen finns inte: {input_dir}", file=sys.stderr)
        return 1

    files = find_matching_files(input_dir)
    if not files:
        print(f'Fel: Inga filer matchade "{FILE_PATTERN}" i {input_dir}.', file=sys.stderr)
        return 1

    created = 0
    failed = 0
    for src_path in files:
        if process_file(src_path, output_dir):
            created += 1
        else:
            failed += 1
        print()

    print("Sammanfattning")
    print(f"  Hittade filer: {len(files)}")
    print(f"  Skapade filer: {created}")
    print(f"  Filer med fel: {failed}")
    print(f"  Utdatamapp: {output_dir}")

    return 0 if failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
